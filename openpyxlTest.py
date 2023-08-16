import openpyxl


def todo():
    from openpyxl import Workbook

    wb = Workbook()

    ws = wb.create_sheet("数据")
    arr = ["userId", "itemId"]
    ws.append(arr)
    wb.remove(wb.worksheets[0])

    row = ["1", "2", "3"]
    colum = ["A", "B", "C", "D"]
    for r in row:
        for c in colum:
            index = c + r
            ws[index] = "hello"
    ws.cell(row=4, column=2, value=10)

    for row in ws.values:
        for value in row:
            print(value)

    wb.save("./files/sample.xlsx")


if __name__ == '__main__':
    todo()
